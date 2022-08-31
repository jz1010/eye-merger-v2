#!/usr/bin/env python

import os
import sys

sys.path.append('./website')

import re
import math
import random
import time
import argparse
from collections import defaultdict
import pi3d
from xml.dom.minidom import parse
from gfxutil import *
from evdev import InputDevice, ecodes
from joystick import joystick_t
from keyboard import keyboard_t
from debug import leak_check
from wearables import wearables_client_t, wearables_server_t
from openpyxl import load_workbook
import datetime
from gecko_json_util import json_db_t

# These need to be globals in order to avoid a memory leak
DISPLAY = pi3d.Display.create(samples=4)
cam    = pi3d.Camera(is_3d=False, at=(0,0,0), eye=(0,0,-1000))
#cam    = pi3d.Camera(is_3d=True, at=(0,0,0), eye=(0,0,-1000)) # Interesting more ball-like

# TODO - choose a shader
shader = pi3d.Shader("uv_light") # default - light source has impact
#shader = pi3d.Shader("uv_flat") # evenly lighted - perhaps this is good
shader_reflect = pi3d.Shader("uv_reflect") # evenly lighted - perhaps this is good
#shader = shader_reflect
#shader = pi3d.Shader("mat_flat") # weird, no lights
#shader = pi3d.Shader("mat_reflect") # weird
#shader = pi3d.Shader("uv_toon") # no video

light  = pi3d.Light(lightpos=(0, -500, -500), lightamb=(0.2, 0.2, 0.2)) # default
#light  = pi3d.Light(lightpos=(0, -500, -500), lightamb=(0.3, 0.3, 0.3)) 
#light  = pi3d.Light(lightpos=(0, -500, -500), lightamb=(0.1, 0.1, 0.1)) # too dim
#light  = pi3d.Light(lightpos=(0, -500, -500), lightamb=(0.8, 0.8, 0.8)) # looks washed out

# Blinking states
BLINK_NONE = 0
BLINK_CLOSING = 1
BLINK_OPENING = 2

class gecko_eye_t(object):
    def __init__(self,debug=False,EYE_SELECT=None):
        self.debug = debug
        self.init_cfg_db()
        self.EYE_SELECT = None
        if EYE_SELECT is not None:
            self.EYE_SELECT = EYE_SELECT
        elif self.EYE_SELECT is None:
            self.EYE_SELECT = os.getenv('EYE_SELECT','hack')

        print('EYE_SELECT: {}'.format(self.EYE_SELECT))
        #raise
    
        self.keyboard = None
        self.joystick = None

        # Periodically check for new devices connected
        self.keyboard_last_retry = 0
        self.joystick_last_retry = 0
        self.test_joystick_cnt = 0

        # Other timer initialization
        self.time_last_joystick_service = 0
        self.nxt_emotion_sec = 0
        self.last_v = 0.5
        self.last_eye_comm_recv = 0
        self.event_holdDuration = None
        self.event_moveDuration = None
        self.event_overrideBlinkDurationClose = None
        self.event_overrideBlinkDurationOpen = None        
        self.event_doBlink = False

        self.eye_range_x = 60.0
        self.eye_range_y = 60.0
        self.eye_half_range_x = self.eye_range_x / 2.0
        self.eye_half_range_y = self.eye_range_y / 2.0                    
    
    
        self.parse_args()
        self.load_constraints()
        
        self.eye_contexts = ['cyclops','hack','dragon']
        self.eye_cache = defaultdict(dict)
        self.init_display()
        
        self.init(self.eye_contexts)

        self.repo_dir = os.getenv('SB',os.getcwd())
        settings_dir = '{}/settings/'.format(self.repo_dir)
        self.json_db = json_db_t(settings_dir=settings_dir)
        self.time_last_poll = datetime.datetime.now()

    def load_constraints(self):
        if self.cfg_db['eye_constraints'] is None:
            return

        print ('loading eye constraints from: {}'.format(self.cfg_db['eye_constraints']))
        wb = load_workbook(filename = self.cfg_db['eye_constraints'])
        wb_sheet_names = wb.get_sheet_names()
        print ('# {}'.format(wb_sheet_names))

        fnames_sclera = [None]
        fnames_iris = [None]
        dir_graphics = '.'
        
        for sheet_name in wb_sheet_names:
            if sheet_name in ['eye']:
                sheet = wb[sheet_name]
                self.eye_constraints = {}
                
                # Row 0 contains schlera file names
                # Col 0 contains iris file names
                # Intersections (row,col) contain boolean if combination is allowed
                for row_idx,row in enumerate(sheet.iter_rows(min_row=0)):
                    rowlen = len(row)
                    #print (rowlen)
                    #raise
                    if row_idx == 0:
                        dir_graphics = row[0].value
                        
                        fnames_sclera += ['{}/{}'.format(dir_graphics,
                                                         row[col_idx].value) \
                                          for col_idx in range(1,rowlen)]

                        continue
                        
                    for col_idx in range(0,rowlen):
                        cell = row[col_idx]
                        val = cell.value

                        if col_idx == 0:
                            fnames_iris.append('{}/{}'.format(dir_graphics,val))
                            continue
                            
                        if type(val) == str:
                            val = val.upper()

                        print ('row_idx: {} col_idx: {}'.format(row_idx,col_idx))
                        key_sclera = fnames_sclera[col_idx]
                        key_iris = fnames_iris[row_idx]
                        key = (key_sclera,key_iris)
                        print ('key: {}'.format(key))
                        if val in ['Y','X']: # combination permitted
                            self.eye_constraints[key] = True
                        else:
                            self.eye_constraints[key] = False

            elif sheet_name in ['wearables']:
                pass
            else:
                print ('Ignoring sheet name: {}'.format(sheet_name))

        print ('fnames_sclera: {}'.format(fnames_sclera))
        print ('fnames_iris: {}'.format(fnames_iris))
        dir_graphics = 'hack_graphics'
        self.hack_scleras = fnames_sclera[1:]
        self.hack_iris = fnames_iris[1:]
        
        #print ('eye_constraints: {}'.format(self.eye_constraints))
        self.fname_sclera = None
        self.fname_iris = None
        
    def parse_args(self):
        self.parser = argparse.ArgumentParser(description="Parse arguments")
        self.parser.add_argument('--demo',
                                 default=self.cfg_db['demo'],
                                 action='store_true',
                                 help='Demo mode (various eye animations)')
        self.parser.add_argument('--playa',
                                 default=self.cfg_db['playa'],
                                 action='store_true',
                                 help='Playa mode (playa eye animations)')
        self.parser.add_argument('--screenshots',
                                 default=self.cfg_db['screenshots'],
                                 action='store_true',
                                 help='Enumerate graphics for screenshots')
        self.parser.add_argument('--joystick_test',
                                 default=self.cfg_db['joystick_test'],
                                 action='store_true',
                                 help='Inject joystick test messages')
        self.parser.add_argument('--joystick_mode',
                                 default=self.cfg_db['joystick_mode'],
                                 action='store',
                                 type=int,
                                 help='Joystick mode of operation - 0:discrete event 1:continuous')
        self.parser.add_argument('--timeout_secs',
                                 default=self.cfg_db['timeout_secs'],
                                 action='store',type=int,
                                 help='Exit application after N seconds')
        self.parser.add_argument('--joystick_service_interval_sec',
                                 default=self.cfg_db['joystick_service_interval_sec'],
                                 action='store',type=float,
                                 help='Poll joystick with period in secs')
        self.parser.add_argument('--move_fast_duration_joystick_sec',
                                 default=self.cfg_db['move_fast_duration_joystick_sec'],
                                 action='store',type=float,
                                 help='Latency of fast eye repositioning in secs')
        self.parser.add_argument('--auto_restart_joystick_interval_sec',
                                 default=self.cfg_db['auto_restart_joystick_interval_sec'],
                                 action='store',type=float,
                                 help='Latency of automode after joystick idle')
        self.parser.add_argument('--autoblink',
                                 default=self.cfg_db['AUTOBLINK'],
                                 action='store',
                                 help='Autoblink of eyelid')
        self.parser.add_argument('--eye_orientation',
                                 default=self.cfg_db['eye_orientation'],
                                 action='store',
                                 help='Eye orientation: right (default), left')
        self.parser.add_argument('--eye_select',
                                 default=self.EYE_SELECT,
                                 action='store',
                                 help='Eye profile selection')
        self.parser.add_argument('--eye_shape',
                                 default=self.cfg_db[self.EYE_SELECT]['eye.shape'],
                                 action='store',
                                 help='Eye shape art file (.svg)')
        self.parser.add_argument('--iris_art',
                                 default=self.cfg_db[self.EYE_SELECT]['iris.art'],        
                                 action='store',
                                 help='Iris art file (.jpg)')
        self.parser.add_argument('--lid_art',
                                 default=self.cfg_db[self.EYE_SELECT]['lid.art'],        
                                 action='store',
                                 help='Lid art file (.png)')
        self.parser.add_argument('--sclera_art',
                                 default=self.cfg_db[self.EYE_SELECT]['sclera.art'],        
                                 action='store',
                                 help='Sclera art file (.png)')
        self.parser.add_argument('--eye_constraints',
                                 default=self.cfg_db['eye_constraints'],
                                 action='store',
                                 help='Eye art combination constraints')
                           
        # Parse the arguments
        args = self.parser.parse_args()

        # Harvest and validate the arguments
        self.cfg_db['AUTOBLINK'] = (int(args.autoblink) != 0)
        self.EYE_SELECT = args.eye_select

        if args.eye_shape not in ["None"]:
            self.cfg_db[self.EYE_SELECT]['eye.shape'] = args.eye_shape
        if args.iris_art not in ["None"]:
            self.cfg_db[self.EYE_SELECT]['iris.art'] = args.iris_art
        if args.lid_art not in ["None"]:            
            self.cfg_db[self.EYE_SELECT]['lid.art'] = args.lid_art
        if args.sclera_art not in ["None"]:                        
            self.cfg_db[self.EYE_SELECT]['sclera.art'] = args.sclera_art

        if args.eye_orientation not in ['left','right']:
            print ('** ERROR: Eye orientation must be either left or right')
            sys.exit(1)

        self.cfg_db['eye_orientation'] = args.eye_orientation
        self.cfg_db['demo'] = args.demo
        self.cfg_db['playa'] = args.playa
        self.cfg_db['screenshots'] = args.screenshots
        self.cfg_db['eye_constraints'] = args.eye_constraints
        self.cfg_db['joystick_mode'] = args.joystick_mode
        self.cfg_db['joystick_service_interval_sec'] = args.joystick_service_interval_sec
        self.cfg_db['move_fast_duration_joystick_sec'] = args.move_fast_duration_joystick_sec
        
        assert (not (self.cfg_db['demo'] and self.cfg_db['playa']))
        
        self.cfg_db['joystick_test'] = args.joystick_test
        self.cfg_db['timeout_secs'] = args.timeout_secs

    
    def init_cfg_db(self):
        self.cfg_db = {
            'demo': False, # Demo mode boolean
            'demo_eye_tenure_secs' : 3,
            'playa' : False, # Playa mode operation
            'joystick_test' : False, # No test messages from joystick
            'joystick_mode' : 0, # 0: Discrete mode, 1: Continuous mode
            'eye_queue_max' : 1, # maximum depth of eye event queue
#            'eye_queue_max' : 5, # maximum depth of eye event queue            
            'screenshots' : False, # Enumerates graphics and take screenshots
            'timeout_secs':None, # 1 hour (60 min * 60 sec)
            'eye_orientation': 'right', # Default right eye orientation
            'JOYSTICK_X_IN': -1,    # Analog input for eye horiz pos (-1 = auto)
            'JOYSTICK_Y_IN': -1,    # Analog input for eye vert position (")
            'PUPIL_IN': -1,    # Analog input for pupil control (-1 = auto)
            'JOYSTICK_X_FLIP': False, # If True, reverse stick X axis
            'JOYSTICK_Y_FLIP': False, # If True, reverse stick Y axis
            'PUPIL_IN_FLIP': False, # If True, reverse reading from PUPIL_IN
            'TRACKING': True,  # If True, eyelid tracks pupil
            #'TRACKING': False,  # If True, eyelid tracks pupil            
            'PUPIL_SMOOTH': 16,    # If > 0, filter input from PUPIL_IN
            'pupil_min': 0.0,   # Lower analog range from PUPIL_IN            
            'pupil_max': 1.0,   # Upper "
            'pupil_normal': 0.5,   # Normal pupil dilation
            'AUTOBLINK' : True,  # If True, eye blinks autonomously
            #'AUTOBLINK'  : False,  # If True, eye blinks autonomously
            'switch_on_blink' : True,
            'blink_angry_duration_close_min_sec' : 0.10, # original: 0.06
            'blink_angry_duration_close_max_sec' : 0.30, # original: 0.12
            'blink_angry_duration_open_min_sec' : 0.06, # original: 0.06
            'blink_angry_duration_open_max_sec' : 0.12, # original: 0.12
            'blink_angry_interval_min_sec' : 1.0, # original: 3.0
            'blink_angry_interval_range_sec' : 3.0, # original: 4.0
            
            'blink_duration_close_min_sec' : 0.5, # original: 0.06
            'blink_duration_close_max_sec' : 1.5, # original: 0.12
            'blink_duration_open_min_sec' : 0.5, # original: 0.06
            'blink_duration_open_max_sec' : 1.5, # original: 0.12
            
            'blink_interval_min_sec' : 15.0, # original: 3.0
            'blink_interval_range_sec' : 4.0, # original: 4.0
            
            'blink_duration_joystickmin_sec' : 0.035, # caused by Joystick trigger
            'blink_duration_joystickmax_sec' : 0.06, # caused by Joystick trigger
            #'joystick_service_interval_sec' : 0.20,
            'joystick_service_interval_sec' : 0.0,            
            
            'pupil_auto_expand_sec' : 12.0,
            'pupil_auto_contract_sec' : 4.0,
            
            'pupil_squint_sec' : 4.0,
            'pupil_dilate_sec' : 4.0,

            'move_angry_duration_min_sec' : 0.075, # original: 0.075
            'move_angry_duration_max_sec' : 0.175, # original: 0.175
            
            'hold_angry_duration_min_sec' : 0.1, # original: 0.1
            'hold_angry_duration_max_sec' : 1.1, # original: 1.1
            
            'move_duration_min_sec' : 7.5, # original: 0.075
            'move_duration_max_sec' : 12.0, # original: 0.175
            
            'hold_duration_min_sec' : 2.1, # original: 0.1
            'hold_duration_max_sec' : 4.0, # original: 1.1

#            'move_fast_duration_joystick_sec' : 0.10, # caused by Joystick position
            'move_fast_duration_joystick_sec' : 0.0, # caused by Joystick position            
            'move_slow_duration_joystick_sec' : 0.50, # caused by Joystick position
            'move_scripted_duration_joystick_sec' : 0.10, # caused by Joystick position            

            'auto_restart_interval_sec' : 4.00, # time after last network msg input auto-mode engages
            'auto_restart_joystick_interval_sec' : 10.00, # time after last joystick input auto-mode engages            
            'emotion_interval_sec' : 8.0, # Interval between next emotion

            'joystick_retry_init_sec' : (2*60), # 5 mins
            'keyboard_retry_init_sec' : (60*60), # 1 hour

            #
            # Eye graphics definitions
            #
            'cyclops': {
		'eye.shape': 'graphics/cyclops-eye.svg',
		'iris.art': 'graphics/iris.jpg',
		'lid.art': 'graphics/lid.png',
		'sclera.art': 'graphics/sclera.png'
		},
            'dragon': {
		'eye.shape': 'graphics/dragon-eye.svg',
		#'iris.art': 'graphics/dragon-iris.jpg',
		'iris.art': 'hack_graphics/Metal_Iris_00145.jpg',                
		'lid.art': 'graphics/lid.png',
		#'sclera.art': 'graphics/dragon-sclera.png'
                'sclera.art': 'hack_graphics/Circuit_sclera_00000.jpg',
            },
            'hack': {
#		'eye.shape': 'graphics/cyclops-eye.svg',                
		'eye.shape': 'hack_graphics/gecko-eye_playa.svg',
#		'eye.shape': 'hack_graphics/gecko-eye_0.svg',                
#		'iris.art': 'hack_graphics/iris.jpg',

                # Trent
		'iris.art': 'hack_graphics/Metal_Iris_00145.jpg',
#		'iris.art': 'hack_graphics/Metal iris animated 1.gif',
#		'iris.art': 'hack_graphics/Metal Iris animated 2_[00168-00288].gif',
#		'iris.art': 'graphics/uv.png',                
#		'iris.art': 'hack_graphics/Organic eye_01081.jpg',                
                
#		'iris.art': 'hack_graphics/dragon-iris.jpg',                
		'lid.art': 'hack_graphics/lid.png',
#		'lid.art': 'graphics/uv.png',                
                
#		'sclera.art': 'hack_graphics/dragon-sclera.png',
#		'sclera.art': 'hack_graphics/dragon-iris.jpg'
#		'sclera.art': 'hack_graphics/gecko_s_eye_by_mchahine_d2en705-fullview.jpg'
#                'sclera.art': 'hack_graphics/leopard-gecko-3381555_960_720.jpg',
#                'sclera.art': 'hack_graphics/Ds4CWFgV4AAlhWK.jpg_large.jpg',
#                'sclera.art': 'hack_graphics/sclera.jpg',

                # Trent
                'sclera.art': 'hack_graphics/Circuit_sclera_00000.jpg',
#		'sclera.art': 'hack_graphics/Organic eye_01081.jpg',                
	    },
            'eye_constraints' : None,
            
            #
            # Network related
            #
            
            # Wearables
            'port_wearables' : 0xDF0D,
            'mcaddr_wearables' : '239.255.223.01',

            # Inter-eye communication
            'port_eyes' : 0xDF0D,
            'mcaddr_eyes' : '239.255.223.02'
        }

        self.hack_eye_shapes = [
	    'graphics/cyclops-eye.svg',
	    'graphics/dragon-eye.svg',            
        ]
        
        self.hack_scleras = [
            'hack_graphics/Circuit_sclera_00000.jpg',
            'hack_graphics/Circuit_sclera_00001.jpg',
            'hack_graphics/Circuit_sclera_00002.jpg',
            'hack_graphics/Circuit_sclera_00003.jpg',
            'hack_graphics/Circuit_sclera_00004.jpg',
            'hack_graphics/Circuit_sclera_00005.jpg',
            'hack_graphics/Circuit_sclera_00006.jpg',
            'hack_graphics/Circuit_sclera_00007.jpg',
            'hack_graphics/Circuit_sclera_00008.jpg',
            'hack_graphics/Circuit_sclera_00009.jpg',
            'hack_graphics/Circuit_sclera_00010.jpg',
            'hack_graphics/Circuit_sclera_00011.jpg',
            'hack_graphics/Circuit_sclera_00012.jpg',
            'hack_graphics/Circuit_sclera_00013.jpg',
            'hack_graphics/Circuit_sclera_00014.jpg',
            'hack_graphics/Circuit_sclera_00015.jpg',
            'hack_graphics/gecko_spiral_eyes_01281.jpg',
            'hack_graphics/gecko_spiral_eyes_01282.jpg',
            'hack_graphics/gecko_spiral_eyes_01283.jpg',
            'hack_graphics/gecko_spiral_eyes_01284.jpg',
            'hack_graphics/gecko_spiral_eyes_01285.jpg',
            'hack_graphics/gecko_spiral_eyes_01286.jpg',
        ]
        self.hack_iris = [
	    'hack_graphics/Metal_Iris_00145.jpg',
	    'hack_graphics/Metal_Iris_00146.jpg',
	    'hack_graphics/Metal_Iris_00147.jpg',
	    'hack_graphics/Metal_Iris_00148.jpg',
	    'hack_graphics/Metal_Iris_00149.jpg',
	    'hack_graphics/Metal_Iris_00150.jpg',
	    'hack_graphics/Metal_Iris_00151.jpg',
	    'hack_graphics/Metal_Iris_00152.jpg',
	    'hack_graphics/Metal_Iris_00153.jpg',
            'hack_graphics/gecko_spiral_eyes_01281.jpg',
            'hack_graphics/gecko_spiral_eyes_01282.jpg',
            'hack_graphics/gecko_spiral_eyes_01283.jpg',
            'hack_graphics/gecko_spiral_eyes_01284.jpg',
            'hack_graphics/gecko_spiral_eyes_01285.jpg',
            'hack_graphics/gecko_spiral_eyes_01286.jpg',
        ]

#        self.hack_iris = [
#            'hack_graphics/gecko spiral eyes_01281.jpg',
#            'hack_graphics/gecko spiral eyes_01282.jpg',
#            'hack_graphics/gecko spiral eyes_01283.jpg',
#            'hack_graphics/gecko spiral eyes_01284.jpg',
#            'hack_graphics/gecko spiral eyes_01285.jpg',
#            'hack_graphics/gecko spiral eyes_01286.jpg',
#        ]

    def switch_eye_context(self,eye_context):
        if eye_context is None:
            return

        print ('switch_eye_context: {}'.format(eye_context))
        if eye_context in ['hack']:
            self.load_textures(eye_context)
            self.iris.set_textures([self.irisMap])
            self.eye.set_textures([self.scleraMap])            
            #self.init_svg(eye_context)
            
        #print (self.eye_cache[eye_context])
        self.vb = self.eye_cache[eye_context]['vb']
        self.pupilMinPts = self.eye_cache[eye_context]['pupilMinPts']
        self.pupilMaxPts = self.eye_cache[eye_context]['pupilMaxPts']
        self.irisPts = self.eye_cache[eye_context]['irisPts']
        self.scleraFrontPts = self.eye_cache[eye_context]['scleraFrontPts']
        self.scleraBackPts = self.eye_cache[eye_context]['scleraBackPts']
        self.upperLidClosedPts = self.eye_cache[eye_context]['upperLidClosedPts']
        self.upperLidOpenPts = self.eye_cache[eye_context]['upperLidOpenPts']
        self.upperLidEdgePts = self.eye_cache[eye_context]['upperLidEdgePts']
        self.lowerLidClosedPts = self.eye_cache[eye_context]['lowerLidClosedPts']
        self.lowerLidOpenPts = self.eye_cache[eye_context]['lowerLidOpenPts']
        self.lowerLidEdgePts = self.eye_cache[eye_context]['lowerLidEdgePts']
        self.irisMap = self.eye_cache[eye_context]['irisMap']
        self.scleraMap = self.eye_cache[eye_context]['scleraMap']
        self.lidMap = self.eye_cache[eye_context]['lidMap']
        self.eye = self.eye_cache[eye_context]['eye']
        self.iris = self.eye_cache[eye_context]['iris']
        self.irisZ = self.eye_cache[eye_context]['irisZ']
        self.upperEyelid = self.eye_cache[eye_context]['upperEyelid']
        self.lowerEyelid = self.eye_cache[eye_context]['lowerEyelid']
        
        if not self.eye_cache[eye_context]['geometry_initialized']:
            self.eye_cache[eye_context]['geometry_initialized'] = True
            self.init_geometry(eye_context)

        #self.init_geometry_iris()
        #self.init_geometry_eyelids()
        #self.init_geometry_sclera()

        self.draw_eye()
        
        return eye_context
        
    def init(self,eye_contexts):
        # Eye context dependent
        
        # Load animations
        if False:
            self.load_animations()
                        
        # Build eyes
        for eye_context in eye_contexts:
            self.init_svg(eye_context)
            self.load_textures(eye_context)
            self.init_geometry(eye_context)
            self.eye_cache[eye_context]['geometry_initialized'] = True
            self.init_globals()
            
        # Set eye context
        #next_eye = 'cyclops'
        #next_eye = 'dragon'
        self.EYE_SELECT = self.switch_eye_context(self.EYE_SELECT)

        # Independent of rendering setup
        self.init_keyboard()
        # self.init_joystick()
        self.init_emotion()
        self.init_wearables()
        self.init_eye_comm()
        
    def find_input_device(self,input_name=None):
        if input_name is None:
            return None

        for i in range(10):
            device_name = '/dev/input/event{}'.format(i)
            if os.path.exists(device_name):
                print ('scanning device: {}'.format(device_name))
                try:
                    device = InputDevice(device_name)
                except:
                    device = None
                if device is not None:
                    print ('device: {}'.format(device))
                    if input_name in ['keyboard'] and \
                       re.search(r'keyboard',device.name,re.IGNORECASE):
                        return device_name
                    elif input_name in ['joystick'] and \
                         ((re.search(r'extreme(\s+)3d',device.name,re.IGNORECASE) or \
                           re.search(r'logitech(\s+)freedom',device.name,re.IGNORECASE))):
                        print('Found joystick device: {}'.format(device_name))
                        return device_name
        return None

    def init_eye_comm(self):
        self.eye_server = wearables_server_t(self.debug,
                                             mcaddr=self.cfg_db['mcaddr_eyes'],
                                             port=self.cfg_db['port_eyes'])
        self.eye_comm_msg_cnt = 0
        self.eye_client = wearables_client_t(self.debug,
                                             mcaddr=self.cfg_db['mcaddr_eyes'],
                                             port=self.cfg_db['port_eyes'])
        
    def init_wearables(self):
        self.wearables_msg_cnt = 0
        self.wearables_client = wearables_client_t(self.debug,
                                                   mcaddr=self.cfg_db['mcaddr_wearables'],
                                                   port=self.cfg_db['port_wearables'])
        
    def init_joystick(self):
        if self.joystick is None:
            now = time.time()
            if now - self.joystick_last_retry <= self.cfg_db['joystick_retry_init_sec']:
                return

            self.joystick_last_retry = now
            
            device_name = self.find_input_device('joystick')
            self.joystick = joystick_t(joystick_dev=device_name,
                                       joystick_mode=self.cfg_db['joystick_mode'])
            self.joystick_polls = 0
            self.joystick_msg_cnt = 0
            self.joystick_range = self.joystick.range_hi
            self.joystick_half_range = self.joystick_range / 2.0
            
            # Set up state kept from events sampled from joystick
            if self.joystick.get_status():
                self.update_eye_events(reset=True)
            else:
                self.joystick = None
            self.debug_joystick_sec = 0
        else:
            if not self.joystick.get_status():
                self.joystick = None

    def init_keyboard(self):
        if self.keyboard is None:
            now = time.time()
            if now - self.keyboard_last_retry <= self.cfg_db['keyboard_retry_init_sec']:
                return

            self.keyboard_last_retry = now

            device_name = self.find_input_device('keyboard')
            self.keyboard = keyboard_t(device_name)
            if not self.keyboard.get_status():
                self.keyboard = None
        else:
            if not self.keyboard.get_status():
                self.keyboard = None
            
    def init_svg(self,eye_context=None):
        # Load SVG file, extract paths & convert to point lists --------------------

        # Thanks Glen Akins for the symmetrical-lidded cyclops eye SVG!
        # Iris & pupil have been scaled down slightly in this version to compensate
        # for how the WorldEye distorts things...looks OK on WorldEye now but might
        # seem small and silly if used with the regular OLED/TFT code.
        dom                    = parse(self.cfg_db[eye_context]['eye.shape'])        
        self.vb                = getViewBox(dom)
        self.pupilMinPts       = getPoints(dom, "pupilMin"      , 32, True , True )
        #self.pupilMinPts       = getPoints(dom, "pupilMin"      , 64, True , True )        
        self.pupilMaxPts       = getPoints(dom, "pupilMax"      , 32, True , True )
        #self.pupilMaxPts       = getPoints(dom, "pupilMax"      , 64, True , True )        
        self.irisPts           = getPoints(dom, "iris"          , 32, True , True )
        self.scleraFrontPts    = getPoints(dom, "scleraFront"   ,  0, False, False)
        self.scleraBackPts     = getPoints(dom, "scleraBack"    ,  0, False, False)
        self.upperLidClosedPts = getPoints(dom, "upperLidClosed", 33, False, True )
        self.upperLidOpenPts   = getPoints(dom, "upperLidOpen"  , 33, False, True )
        self.upperLidEdgePts   = getPoints(dom, "upperLidEdge"  , 33, False, False)
        self.lowerLidClosedPts = getPoints(dom, "lowerLidClosed", 33, False, False)
        self.lowerLidOpenPts   = getPoints(dom, "lowerLidOpen"  , 33, False, False)
        self.lowerLidEdgePts   = getPoints(dom, "lowerLidEdge"  , 33, False, False)

        if eye_context is not None:
            self.eye_cache[eye_context]['vb'] = self.vb
            self.eye_cache[eye_context]['pupilMinPts'] = self.pupilMinPts
            self.eye_cache[eye_context]['pupilMaxPts'] = self.pupilMaxPts
            self.eye_cache[eye_context]['irisPts'] = self.irisPts
            self.eye_cache[eye_context]['scleraFrontPts'] = self.scleraFrontPts
            self.eye_cache[eye_context]['scleraBackPts'] = self.scleraBackPts
            self.eye_cache[eye_context]['upperLidClosedPts'] = self.upperLidClosedPts
            self.eye_cache[eye_context]['upperLidOpenPts'] = self.upperLidOpenPts
            self.eye_cache[eye_context]['upperLidEdgePts'] = self.upperLidEdgePts
            self.eye_cache[eye_context]['lowerLidClosedPts'] = self.lowerLidClosedPts
            self.eye_cache[eye_context]['lowerLidOpenPts'] = self.lowerLidOpenPts
            self.eye_cache[eye_context]['lowerLidEdgePts'] = self.lowerLidEdgePts            

    def init_display(self):
        global DISPLAY
        self.DISPLAY = DISPLAY
        self.DISPLAY.set_background(0, 0, 0, 1) # r,g,b,alpha
        #self.DISPLAY.set_background(0, 128, 0, 1) # green
        #self.DISPLAY.set_background(0, 100, 0, 1) # darkgreen
        #self.DISPLAY.set_background(10, 50, 4, 0) # experimental

        # eyeRadius is the size, in pixels, at which the whole eye will be rendered.
        if self.DISPLAY.width <= (self.DISPLAY.height * 2):
            # For WorldEye, eye size is -almost- full screen height
            if False:
                self.eyeRadius   = self.DISPLAY.height / 2.1
            else:
                self.eyeRadius   = self.DISPLAY.height / 1.7 #1.6 eye spills off screen
        else:
            self.eyeRadius   = self.DISPLAY.height * 2.0 / 5.0

        # A 2D camera is used, mostly to allow for pixel-accurate eye placement,
        # but also because perspective isn't really helpful or needed here, and
        # also this allows eyelids to be handled somewhat easily as 2D planes.
        # Line of sight is down Z axis, allowing conventional X/Y cartesion
        # coords for 2D positions.
        global cam, shader, light
        self.cam = cam
        self.shader = shader
        self.light = light
                

    def constrained_random_eye(self):
        done = False
        while not done:
            fname_iris = random.choice(self.hack_iris)
            fname_sclera = random.choice(self.hack_scleras)
            if fname_iris == self.fname_iris and \
               fname_sclera == self.fname_sclera:
                continue
            
            key = (fname_sclera,fname_iris)
            if self.cfg_db['eye_constraints'] is None:
                done = True
                continue
            
            done = self.eye_constraints[key]

        #fname_eye_shape = random.choice(self.hack_eye_shapes)
        print ('random_eye: {}, {}'.format(
            fname_sclera,
            fname_iris))
        
        return (fname_sclera,fname_iris)

    def load_animations(self,defer_loading=False,eye_context=None):
        dir = 'hack_graphics/animations/sclera'
        from os import listdir
        from os.path import isfile, join
        flist = ['{}'.format(join(dir,f)) for f in listdir(dir) if isfile(join(dir, f))]
        self.animation_frame = 0
        self.animations = []
        for fname in sorted(flist):
            print ('anim fname: {}'.format(fname))
            animation = pi3d.Texture(fname,
                                     mipmap=False, # True, doesn't look as good
                                     defer=defer_loading,                                      
                                     filter=pi3d.GL_LINEAR,
                                     blend=True)
            self.animations.append(animation)
        
    def load_textures(self,eye_context=None):
        # Load texture maps --------------------------------------------------------

        defer_loading = False # Pre-cache textures at setup
        if eye_context in ['hack']:
            if self.cfg_db['playa']:
                (self.fname_sclera,
                 self.fname_iris) = self.constrained_random_eye()
            else:
                self.fname_iris = self.cfg_db[eye_context]['iris.art']
                self.fname_sclera = self.cfg_db[eye_context]['sclera.art']
                pass # do not change art assets in eye
        else:
            self.fname_iris = self.cfg_db[eye_context]['iris.art']
            self.fname_sclera = self.cfg_db[eye_context]['sclera.art']

        print ('fname_sclera: {}'.format(self.fname_sclera))            
        print ('fname_iris: {}'.format(self.fname_iris))
        #print ('fname_eye_shape: {}'.format(self.cfg_db[eye_context]['eye.shape']))

        self.irisMap = pi3d.Texture(self.fname_iris,
                                    mipmap=False, # True, doesn't look as good
                                    defer=defer_loading,
                                    filter=pi3d.GL_LINEAR
#                                   filter=pi3d.GL_NEAREST  # Doesn't look as good
        )
        self.scleraMap = pi3d.Texture(self.fname_sclera,
                                      mipmap=False, # True, doesn't look as good
                                      defer=defer_loading,                                      
                                      filter=pi3d.GL_LINEAR,
#                                     filter=pi3d.GL_NEAREST, # Doesn't look as good
                                      blend=True
#                                     blend=False # No apparent change
        )
        self.lidMap = pi3d.Texture(self.cfg_db[eye_context]['lid.art'],
                                   mipmap=False, # True, doesn't look as good
                                   filter=pi3d.GL_LINEAR,
                                   defer=defer_loading,                                   
#                                  filter=pi3d.GL_NEAREST, # Doesn't look as good
                                   blend=True
#                                  blend=False # No apparent change
        )
        # U/V map may be useful for debugging texture placement; not normally used
        #self.uvMap = pi3d.Texture(self.cfg_db[self.EYE_SELECT]['uv.art'], mipmap=False,
        #                          filter=pi3d.GL_LINEAR, blend=False, m_repeat=True)

        if eye_context is not None:
            self.eye_cache[eye_context]['irisMap'] = self.irisMap
            self.eye_cache[eye_context]['scleraMap'] = self.scleraMap
            self.eye_cache[eye_context]['lidMap'] = self.lidMap
            #self.eye_cache[eye_context]['uvMap'] = self.uvMap

    def init_geometry_iris(self,eye_context=None):
        # Generate initial iris mesh; vertex elements will get replaced on
        # a per-frame basis in the main loop, this just sets up textures, etc.
        if self.cfg_db['eye_orientation'] in ['right']:        
            self.iris = meshInit(32, 4, True, 0, 0.5/self.irisMap.iy, False)
        elif self.cfg_db['eye_orientation'] in ['left']:
            self.iris = meshInit(32, 4, True, 0.5, 0.5/self.irisMap.iy, False)
        else:
            raise
            
        self.iris.set_textures([self.irisMap])
        self.iris.set_shader(self.shader)
        #self.iris.set_shader(shader_reflect)
        self.irisZ = zangle(self.irisPts, self.eyeRadius)[0] * 0.99 # Get iris Z depth, for later

        if eye_context is not None:
            self.eye_cache[eye_context]['iris'] = self.iris
            self.eye_cache[eye_context]['irisZ'] = self.irisZ
        
    def init_geometry_eyelids(self,eye_context=None):
        # Eyelid meshes are likewise temporary; texture coordinates are
        # assigned here but geometry is dynamically regenerated in main loop.
        self.upperEyelid = meshInit(33, 5, False, 0, 0.5/self.lidMap.iy, True)
        #self.upperEyelid = meshInit(40, 7, False, 0, 0.5/self.lidMap.iy, True)        
        self.upperEyelid.set_textures([self.lidMap])
        self.upperEyelid.set_shader(self.shader)
        #self.upperEyelid.set_shader(shader_reflect)
        self.lowerEyelid = meshInit(33, 5, False, 0, 0.5/self.lidMap.iy, True)
        self.lowerEyelid.set_textures([self.lidMap])
        self.lowerEyelid.set_shader(self.shader)
        #self.lowerEyelid.set_shader(shader_reflect)

        if eye_context is not None:
            self.eye_cache[eye_context]['upperEyelid'] = self.upperEyelid
            self.eye_cache[eye_context]['lowerEyelid'] = self.lowerEyelid

    def init_geometry_sclera(self,eye_context=None):
        # Generate sclera for eye...start with a 2D shape for lathing...
        angle1 = zangle(self.scleraFrontPts, self.eyeRadius)[1] # Sclera front angle
        angle2 = zangle(self.scleraBackPts , self.eyeRadius)[1] # " back angle
        aRange = 180 - angle1 - angle2
        pts    = []
        steps = 24
        #steps = 12
        for i in range(steps):
            ca, sa = pi3d.Utility.from_polar((90 - angle1) - aRange * i / (steps-1))
            pts.append((ca * self.eyeRadius, sa * self.eyeRadius))

        #self.eye = pi3d.Lathe(path=pts, sides=16) # artifacts
        #self.eye = pi3d.Lathe(path=pts, sides=32) 
        #self.eye = pi3d.Lathe(path=pts, sides=64) # original
        #self.eye = pi3d.Lathe(path=pts, sides=128)        
        self.eye = pi3d.Lathe(path=pts, sides=256)
        #self.eye = pi3d.Lathe(path=pts, sides=512)
        #self.eye = pi3d.Lathe(path=pts, sides=1024)
        #self.eye = pi3d.Lathe(path=pts, sides=2048)        
        self.eye.set_textures([self.scleraMap])
        self.eye.set_shader(self.shader)
        #self.eye.set_shader(shader_reflect)
        if self.cfg_db['eye_orientation'] in ['right']:
            reAxis(self.eye, 0.0)
        elif self.cfg_db['eye_orientation'] in ['left']:
            reAxis(self.eye, 0.5)
        else:
            raise

        if eye_context is not None:
            self.eye_cache[eye_context]['eye'] = self.eye

        
    def init_geometry(self,eye_context=None):
        # Initialize static geometry -----------------------------------------------

        # Transform point lists to eye dimensions
        offset_x = 0.0
        offset_y = 0.0
        scalePoints(self.pupilMinPts      , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.pupilMaxPts      , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.irisPts          , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.scleraFrontPts   , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.scleraBackPts    , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.upperLidClosedPts, self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.upperLidOpenPts  , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.upperLidEdgePts  , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.lowerLidClosedPts, self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.lowerLidOpenPts  , self.vb, offset_x, offset_y, self.eyeRadius)
        scalePoints(self.lowerLidEdgePts  , self.vb, offset_x, offset_y, self.eyeRadius)

        # Regenerating flexible object geometry (such as eyelids during blinks, or
        # iris during pupil dilation) is CPU intensive, can noticably slow things
        # down, especially on single-core boards.  To reduce this load somewhat,
        # determine a size change threshold below which regeneration will not occur;
        # roughly equal to 1/2 pixel, since 2x2 area sampling is used.

        # Determine change in pupil size to trigger iris geometry regen
        irisRegenThreshold = 0.0
        a = pointsBounds(self.pupilMinPts) # Bounds of pupil at min size (in pixels)
        b = pointsBounds(self.pupilMaxPts) # " at max size
        maxDist = max(abs(a[0] - b[0]), abs(a[1] - b[1]), # Determine distance of max
                      abs(a[2] - b[2]), abs(a[3] - b[3])) # variance around each edge
        # maxDist is motion range in pixels as pupil scales between 0.0 and 1.0.
        # 1.0 / maxDist is one pixel's worth of scale range.  Need 1/2 that...
        if maxDist > 0: self.irisRegenThreshold = 0.5 / maxDist

        # Determine change in eyelid values needed to trigger geometry regen.
        # This is done a little differently than the pupils...instead of bounds,
        # the distance between the middle points of the open and closed eyelid
        # paths is evaluated, then similar 1/2 pixel threshold is determined.
        self.upperLidRegenThreshold = 0.0
        self.lowerLidRegenThreshold = 0.0
        p1 = self.upperLidOpenPts[len(self.upperLidOpenPts) / 2]
        p2 = self.upperLidClosedPts[len(self.upperLidClosedPts) / 2]
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        d  = dx * dx + dy * dy
        if d > 0: self.upperLidRegenThreshold = 0.5 / math.sqrt(d)
        p1 = self.lowerLidOpenPts[len(self.lowerLidOpenPts) / 2]
        p2 = self.lowerLidClosedPts[len(self.lowerLidClosedPts) / 2]
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        d  = dx * dx + dy * dy
        if d > 0: self.lowerLidRegenThreshold = 0.5 / math.sqrt(d)

        self.init_geometry_iris(eye_context)
        self.init_geometry_eyelids(eye_context)
        self.init_geometry_sclera(eye_context)

        
    def init_globals(self):
        # Init global stuff --------------------------------------------------------
        
        self.startX       = random.uniform(-30.0, 30.0)
        n = math.sqrt(900.0 - self.startX * self.startX)
        self.startY       = random.uniform(-n, n)
        self.destX        = self.startX
        self.destY        = self.startY
        self.curX         = self.startX
        self.curY         = self.startY
        self.moveDuration = random.uniform(self.cfg_db['move_duration_min_sec'],
                                           self.cfg_db['move_duration_max_sec'])
        self.holdDuration = random.uniform(self.cfg_db['hold_duration_min_sec'],
                                           self.cfg_db['hold_duration_max_sec'])
        self.move_startTime    = 0.0
        self.isMoving     = False

        self.frame_cnt        = 0
        self.beginningTime = time.time()

        self.eye.positionX(0.0)
        self.iris.positionX(0.0)
        self.upperEyelid.positionX(0.0)
        self.upperEyelid.positionZ(-self.eyeRadius - 42)
        self.lowerEyelid.positionX(0.0)
        self.lowerEyelid.positionZ(-self.eyeRadius - 42)

        self.currentPupilScale  =  0.5
        self.prevPupilScale     = -1.0 # Force regen on first frame
        self.prevUpperLidWeight = 0.5
        self.prevLowerLidWeight = 0.5
        self.prevUpperLidPts    = pointsInterp(self.upperLidOpenPts, self.upperLidClosedPts, 0.5)
        self.prevLowerLidPts    = pointsInterp(self.lowerLidOpenPts, self.lowerLidClosedPts, 0.5)
        
        self.ruRegen = True
        self.rlRegen = True

        self.timeOfLastBlink = 0.0
        self.timeToNextBlink = 1.0
        self.override_timeToNextBlink = None
        self.blinkState      = BLINK_NONE
        self.blinkDuration   = 0.1
        self.blinkStartTime  = 0

        # Eye context switch setup
        self.eye_switch_pending = False
        self.eye_switch_next = None
        
        self.trackingPos = 0.3        

        self.update_eye_events(reset=True)
        test_time = self.cfg_db['move_scripted_duration_joystick_sec']
        self.test_eye_events = [('eye_right',test_time),
                                ('eye_northeast',test_time),
                                ('eye_up',test_time),
                                ('eye_northwest',test_time),
                                ('eye_left',test_time),
                                ('eye_southwest',test_time),
                                ('eye_down',test_time),
                                ('eye_southeast',test_time)]
        
    def split(self, # Recursive simulated pupil response when no analog sensor
              startValue, # Pupil scale starting value (0.0 to 1.0)
              endValue,   # Pupil scale ending value (")
              duration,   # Start-to-end time, floating-point seconds
              range):     # +/- random pupil scale at midpoint
        
        do_exit = False
	startTime = time.time()
	if range >= 0.125: # Limit subdvision count, because recursion
	    duration *= 0.5 # Split time & range in half for subdivision,
	    range    *= 0.5 # then pick random center point within range:
	    midValue  = ((startValue + endValue - range) * 0.5 +
		         random.uniform(0.0, range))
	    do_exit |= self.split(startValue, midValue, duration, range)
            if not do_exit:
		do_exit |= self.split(midValue  , endValue, duration, range)
            self.init_keyboard()
	else: # No more subdivisons, do iris motion...
	    dv = endValue - startValue
	    while not do_exit:
                now_sec = time.time()
		dt = now_sec - startTime
		if dt >= duration: break
                if self.pupil_event_queued: break
                if self.eye_context_next is not None: break                        
		v = startValue + dv * dt / duration
		if   v < self.cfg_db['pupil_min']: v = self.cfg_db['pupil_min']
		elif v > self.cfg_db['pupil_max']: v = self.cfg_db['pupil_max']
		self.frame(v) # Draw frame w/interim pupil scale value
                self.do_wearables()
                self.do_eye_comm()                        
                self.do_joystick()
                do_exit |= self.keyboard_sample()

                if int(now_sec - self.last_eye_art_sec) > self.cfg_db['demo_eye_tenure_secs']:
                    self.last_eye_art_sec = now_sec
                    if (self.cfg_db['demo'] or self.cfg_db['playa']):
                        next_eye = self.random_next_eye()
                        if self.cfg_db['switch_on_blink']:
                            if not self.eye_switch_pending:
                                self.eye_switch_pending = True
                                self.eye_switch_next = next_eye
                        else:
                            self.EYE_SELECT = self.switch_eye_context(next_eye)
                    else:
                        self.EYE_SELECT = self.switch_eye_context(self.EYE_SELECT)
                        self.eye_switch_pending = False

        return do_exit

    def draw_eye(self):
	#convergence = 2.0
        convergence = 0.0        
        if self.cfg_db['eye_orientation'] in ['right']:
	    self.iris.rotateToX(self.curY)
	    self.iris.rotateToY(self.curX - convergence)
	    self.iris.draw()
	    self.eye.rotateToX(self.curY - convergence)
	    self.eye.rotateToY(self.curX)
        elif self.cfg_db['eye_orientation'] in ['left']:
	    self.iris.rotateToX(self.curY)
	    self.iris.rotateToY(self.curX + convergence)
	    self.iris.draw()
	    self.eye.rotateToX(self.curY)
	    self.eye.rotateToY(self.curX + convergence)
        else:
            raise


        if False:
            anim_texture = self.animations[self.animation_frame]
            self.animation_frame += 1
            self.animation_frame %= len(self.animations)
            self.eye.set_textures([anim_texture])
            #self.eye.set_textures(self.animations)
            #self.eye.set_offset(self.animation_frame)
        
	self.eye.draw()
        self.upperEyelid.draw()
        self.lowerEyelid.draw()

    def check_same_event(self,prev_event,event):
        prev_event_opt = None
        event_opt = None
        if type(event) is tuple and type(prev_event) is tuple:
            if event == prev_event:
                pass
            else:
                eye_look_direction = event[0]
                eye_movement_rate = event[1]
                if len(event) > 2:
                    event_remainder = event[2:]
                    
                prev_eye_look_direction = prev_event[0]
                prev_eye_movement_rate = prev_event[1]
                if len(prev_event) > 2:
                    prev_event_remainder = prev_event[2:]

                if eye_look_direction in ['eye_goto'] and \
                   eye_look_direction == prev_eye_look_direction and \
                   eye_movement_rate in ['fast'] and \
                   eye_movement_rate == prev_eye_movement_rate and \
                   event_remainder is not None and \
                   event_remainder != prev_event_remainder:
                    prev_event_opt = None
                    event_opt = event
                elif eye_look_direction == prev_eye_look_direction:
                    assert('fast' in [eye_movement_rate,prev_eye_movement_rate])
                    prev_event_opt = None
                    event_opt = event
                else:
                    prev_event_opt = None
                    event_opt = event
        else:
            if event != prev_event:
                prev_event_opt = None
                event_opt = event

        if self.debug:
            print ('event_opt: {}'.format(event_opt))
        
        return (prev_event_opt,event_opt)
        
    def opt_eye_event_queue(self):
        if True:
            print ('Len(frame_event_queue): {}'.format(len(self.eye_event_queue)))
            
        if self.debug:
            print ('frame_event_queue: {}'.format(self.eye_event_queue))
        events_opt = []
        if len(self.eye_event_queue) > 1:
            #prev_event = self.eye_event_queue[0]
            prev_event = None
            for event in self.eye_event_queue:
                if self.debug:
                    print ('comparing: {} with {}'.format(prev_event,event))
                (prev_event_opt,event_opt) = self.check_same_event(prev_event,event)
                
                if prev_event_opt is not None:
                    events_opt.append(prev_event_opt)
                    
                if event_opt is not None:
                    events_opt.append(event_opt)
                    prev_event = event_opt

            if len(events_opt) == 0:
                events_opt.append(prev_event)
        else:
            events_opt = self.eye_event_queue

        if self.debug:
            print ('pruned_events: {}'.format(events_opt))
        
        return events_opt

    # Generate one frame of imagery
    def frame(self,p):
	self.DISPLAY.loop_running()

	now_sec = time.time()
	dt  = now_sec - self.move_startTime

	self.frame_cnt += 1
        if (self.frame_cnt % 1000) == 0:
            frame_rate = float(self.frame_cnt) / float(now_sec - self.run_start_time)
            if False:
                print ('frame rate: {}'.format(frame_rate))
            
#	if(now_sec > beginningTime):
#		print(frames/(now_sec-beginningTime))

	if self.cfg_db['JOYSTICK_X_IN'] >= 0 and self.cfg_db['JOYSTICK_Y_IN'] >= 0:
            raise
            # Eye position from analog inputs
            self.curX = adcValue[self.cfg_db['JOYSTICK_X_IN']]
            self.curY = adcValue[self.cfg_db['JOYSTICK_Y_IN']]
            if self.cfg_db['JOYSTICK_X_FLIP']: self.curX = 1.0 - self.curX
            if self.cfg_db['JOYSTICK_Y_FLIP']: self.curY = 1.0 - self.curY
            self.curX = -30.0 + self.curX * 60.0
            self.curY = -30.0 + self.curY * 60.0
	else :
            if self.isMoving: # Movement/re-positioning is ongoing
                if dt <= self.moveDuration:
                    scale        = (now_sec - self.move_startTime) / self.moveDuration
                    # Ease in/out curve: 3*t^2-2*t^3
                    scale = 3.0 * scale * scale - 2.0 * scale * scale * scale
                    self.curX         = self.startX + (self.destX - self.startX) * scale
                    self.curY         = self.startY + (self.destY - self.startY) * scale
                else:
                    self.startX       = self.destX
                    self.startY       = self.destY
                    self.curX         = self.destX
                    self.curY         = self.destY
                    if self.event_holdDuration is not None:
                        self.holdDuration = self.event_holdDuration
                        #self.event_holdDuration = None
                    else:
                        self.holdDuration = random.uniform(self.cfg_db['hold_duration_min_sec'],
                                                           self.cfg_db['hold_duration_max_sec'])
                    self.move_startTime    = now_sec
                    self.isMoving     = False                    
            elif self.eye_event_queued() and \
               now_sec >= (self.time_last_joystick_service + \
                          self.cfg_db['joystick_service_interval_sec']):
               # Joystick control has next priority
                self.time_last_joystick_service = now_sec
                self.eye_event_prev = self.eye_event
                if self.cfg_db['joystick_mode'] not in [1]:
                    self.eye_event_queue = self.opt_eye_event_queue()
                self.eye_event = self.eye_event_queue.pop(0)
                if False:
                    print ('frame event: {}'.format(self.eye_event))
                
                if type(self.eye_event) is tuple:
                    eye_look_direction = self.eye_event[0]
                    eye_movement_rate = self.eye_event[1]
                else:
                    eye_look_direction = self.eye_event
                    eye_movement_rate = None

                # Cause eye direction
                if eye_look_direction in ['eye_goto']:
                    #print (self.eye_event)
                    x = self.eye_event[2]
                    y = self.eye_event[3]
                    if x is not None:
                        if x >= (self.joystick_range / 2.0):
                            self.destX = - (self.eye_half_range_x - \
                                            ((self.joystick_range - x) / self.joystick_half_range) * \
                                            self.eye_half_range_x)
                        else:
                            self.destX = ((self.joystick_half_range - x) / self.joystick_half_range) * \
                                self.eye_half_range_x
                    if y is not None:
                        if y >= (self.joystick_range / 2.0):
                            self.destY = - (self.eye_half_range_y - \
                                            ((self.joystick_range - y) / self.joystick_half_range) * \
                                            self.eye_half_range_y)
                        else:
                            self.destY = ((self.joystick_half_range - y) / self.joystick_half_range) * \
                                self.eye_half_range_y
                    #print ('destX: {} destY: {}'.format(self.destX,self.destY))
                elif eye_look_direction in ['eye_up']:
                    self.destX = 0.0
                    n = math.sqrt(900.0 - self.destX * self.destX)
                    self.destY = n
                elif eye_look_direction in ['eye_down']:
                    self.destX = 0.0                    
                    n = math.sqrt(900.0 - self.destX * self.destX)
                    self.destY = -n
                elif eye_look_direction in ['eye_left']:
                    self.destX = 30.0
                    self.destY = 0.0
                elif eye_look_direction in ['eye_right']:
                    self.destX = -30.0
                    self.destY = 0.0
                elif eye_look_direction in ['eye_center']:
                    self.destX = 0.0
                    self.destY = 0.0
                elif eye_look_direction in ['eye_northeast']:
                    self.destX = -30.0
                    n = math.sqrt(900.0)
                    self.destY = n
                elif eye_look_direction in ['eye_northwest']:
                    self.destX = 30.0
                    n = math.sqrt(900.0)
                    self.destY = n
                elif eye_look_direction in ['eye_southeast']:
                    self.destX = -30.0
                    n = math.sqrt(900.0)
                    self.destY = -n
                elif eye_look_direction in ['eye_southwest']:
                    self.destX = 30.0
                    n = math.sqrt(900.0)
                    self.destY = -n
                else:
                    print ('ERROR: Unhandled eye_look_direciton: {}'.format(eye_look_direction))
                    raise

                if eye_movement_rate in ['slow']: 
                    self.event_moveDuration = self.cfg_db['move_slow_duration_joystick_sec']
                    # limit magnitude of movement                    
                    self.destX /= 2
                    self.destY /= 2
                elif eye_movement_rate in ['scripted']:
                    self.event_moveDuration = self.cfg_db['move_scripted_duration_joystick_sec']
                elif eye_movement_rate in ['fast']:
                    self.event_moveDuration = self.cfg_db['move_fast_duration_joystick_sec']
                elif eye_movement_rate in ['slow']:
                    self.event_moveDuration = self.cfg_db['move_slow_duration_joystick_sec']
                else:
                    pass
                
                if self.event_moveDuration is not None:
                    self.moveDuration = self.event_moveDuration
                    self.event_moveDuration = None
                self.move_startTime    = now_sec
                self.isMoving     = True
                
                #self.update_eye_events(reset=True)
            else:
                auto_eye = False
                if self.joystick is None:
                    if now_sec >= self.cfg_db['auto_restart_interval_sec'] + \
                       self.last_eye_comm_recv:
                        auto_eye = True
                else: # joystick is attached
                    if now_sec >= self.cfg_db['auto_restart_joystick_interval_sec'] + \
                       self.joystick.get_last_joystick_time():
                        # resume auto-eye animation if joystick has been idle
                        auto_eye = True

                if auto_eye:
                    if dt >= self.holdDuration:
                        if self.fsm_angry is None:
                            self.destX = random.uniform(-30.0, 30.0)
                            n = math.sqrt(900.0 - self.destX * self.destX)
                            self.destY = random.uniform(-n, n)
                        
                        if self.event_moveDuration is not None:
                            self.moveDuration = self.event_moveDuration
                            self.event_moveDuration = None
                        else:
                            self.moveDuration = random.uniform(
                                self.cfg_db['move_duration_min_sec'],
                                self.cfg_db['move_duration_max_sec'])
                        self.move_startTime = now_sec
                        self.isMoving = True

	# Regenerate iris geometry only if size changed by >= 1/2 pixel
	if abs(p - self.prevPupilScale) >= self.irisRegenThreshold:
		# Interpolate points between min and max pupil sizes
		interPupil = pointsInterp(self.pupilMinPts, self.pupilMaxPts, p)
		# Generate mesh between interpolated pupil and iris bounds
		mesh = pointsMesh(None, interPupil, self.irisPts, 4, -self.irisZ, True)
		self.iris.re_init(pts=mesh)
		self.prevPupilScale = p

	# Eyelid WIP

	if self.event_doBlink or \
           (self.cfg_db['AUTOBLINK'] and \
            (now_sec - self.timeOfLastBlink) >= self.timeToNextBlink):
	    # Similar to movement, eye blinks are slower in this version
	    self.timeOfLastBlink = now_sec
            if self.event_overrideBlinkDurationClose is not None:
                duration = self.event_overrideBlinkDurationClose
                self.event_overrideBlinkDurationClose = None
            else:
		duration = random.uniform(self.cfg_db['blink_duration_close_min_sec'],
                                          self.cfg_db['blink_duration_close_max_sec'])
	    if self.blinkState != BLINK_CLOSING: # infer BLINK_NONE (or BLINK_OPENING?)
		self.blinkState = BLINK_CLOSING 
		self.blinkStartTime = now_sec
		self.blinkDuration = duration
            if self.override_timeToNextBlink is not None:
                self.timeToNextBlink = self.override_timeToNextBlink
                self.override_timeToNextBlink = None                
            else:
                self.timeToNextBlink = duration * self.cfg_db['blink_interval_min_sec'] + \
                    random.uniform(0.0,self.cfg_db['blink_interval_range_sec'])
            self.event_doBlink = False

	if self.blinkState: # Eye currently winking/blinking (BLINK_CLOSING or BLINK_OPENING)
	    # Check if blink time has elapsed...
	    if (now_sec - self.blinkStartTime) >= self.blinkDuration:
		# Yes...increment blink state, unless...
		if (self.blinkState == BLINK_CLOSING and # Enblinking and...
                    self.event_blink == 0):
		    # Don't advance yet; eye is held closed
                    pass
		else:
		    self.blinkState += 1
		    if self.blinkState > BLINK_OPENING:
			self.blinkState = BLINK_NONE # NOBLINK
		    else: # infer BLINK_OPENING
                        assert (self.blinkState == BLINK_OPENING)
                        if self.eye_switch_pending:
                            self.eye_switch_pending = False
                            self.EYE_SELECT = self.switch_eye_context(self.eye_switch_next)
                            self.eye_switch_next = None
                        if self.event_overrideBlinkDurationOpen is not None:
                            duration = self.event_overrideBlinkDurationOpen
                            #self.event_overrideBlinkDurationOpen = None
                        else:
		            duration = \
                                random.uniform(self.cfg_db['blink_duration_open_min_sec'],
                                               self.cfg_db['blink_duration_open_max_sec'])
		            self.blinkDuration = duration
			    self.blinkStartTime = now_sec
	else: # infer BLINK_NONE
            if self.event_blink == 0:
                self.blinkState     = BLINK_CLOSING
                self.blinkStartTime = now_sec
                self.blinkDuration  = random.uniform(self.cfg_db['blink_duration_joystickmin_sec'],
                                                     self.cfg_db['blink_duration_joystickmax_sec'])

	if self.cfg_db['TRACKING']:
		# 0 = fully up, 1 = fully down
		n = 0.5 - self.curY / 70.0
		if   n < 0.0: n = 0.0
		elif n > 1.0: n = 1.0
		self.trackingPos = (self.trackingPos * 3.0 + n) * 0.25

	if self.blinkState: # blink opening/closing
		n = (now_sec - self.blinkStartTime) / self.blinkDuration
		if n > 1.0:
                    n = 1.0
		if self.blinkState == BLINK_OPENING:
                    n = 1.0 - n
	else: # infer BLINK_NONE, Not blinking
		n = 0.0
        self.newUpperLidWeight = self.trackingPos + (n * (1.0 - self.trackingPos))
	self.newLowerLidWeight = (1.0 - self.trackingPos) + (n * self.trackingPos)

        if self.cfg_db['eye_orientation'] in ['right']:                
            flip = True
        elif self.cfg_db['eye_orientation'] in ['left']:            
            flip = False
        else:
            raise
        
	if (self.ruRegen or \
            (abs(self.newUpperLidWeight - self.prevUpperLidWeight) >= \
             self.upperLidRegenThreshold)):
            self.newUpperLidPts = pointsInterp(self.upperLidOpenPts,
                                               self.upperLidClosedPts,
                                               self.newUpperLidWeight)
            if self.newUpperLidWeight > self.prevUpperLidWeight:
                self.upperEyelid.re_init(
                    pts=pointsMesh(
                    self.upperLidEdgePts,
                    self.prevUpperLidPts,
                    self.newUpperLidPts, 5, 0, False, flip))
            else:
                self.upperEyelid.re_init(
                    pts=pointsMesh(
                    self.upperLidEdgePts,
                    self.newUpperLidPts,
                    self.prevUpperLidPts, 5, 0, False, flip))
            self.prevUpperLidWeight = self.newUpperLidWeight
            self.prevUpperLidPts    = self.newUpperLidPts
            self.ruRegen = True
	else:
            self.ruRegen = False

	if (self.rlRegen or \
            (abs(self.newLowerLidWeight - self.prevLowerLidWeight) >= \
             self.lowerLidRegenThreshold)):
            self.newLowerLidPts = pointsInterp(self.lowerLidOpenPts,
                                               self.lowerLidClosedPts,
                                               self.newLowerLidWeight)
            if self.newLowerLidWeight > self.prevLowerLidWeight:
                self.lowerEyelid.re_init(
                    pts=pointsMesh(
                    self.lowerLidEdgePts,
                    self.prevLowerLidPts,
                    self.newLowerLidPts, 5, 0, False, flip))
            else:
                self.lowerEyelid.re_init(
                    pts=pointsMesh(
                    self.lowerLidEdgePts,
                    self.newLowerLidPts,
                    self.prevLowerLidPts, 5, 0, False, flip))
            self.prevLowerLidWeight = self.newLowerLidWeight
            self.prevLowerLidPts    = self.newLowerLidPts
            self.rlRegen = True
	else:
            self.rlRegen = False

	# Draw eye
        self.draw_eye()

    def keyboard_sample(self):
        self.init_keyboard()
        if self.keyboard is not None:
            events = self.keyboard.sample()
            for event in events:
                print ('event: {}'.format(event))
                if event.type == ecodes.EV_KEY:
                    print ('keyboard event: {}'.format(event))
                    if event.code == 1 and event.value == 1: # Escape key
                        return True
                elif True: # other keys here
                    pass
                else:
                    pass
                    
        return False
    
    def update_eye_events(self,reset=False):
        if reset:
            self.eye_event = None
            self.eye_event_queue = []
            self.eye_event_prev = None
            self.eye_context_next = None
            
            self.pupil_event_queued = False
            self.pupil_event_last = None
            self.event_blink = 1
            
    def eye_event_queued(self):
        return len(self.eye_event_queue) > 0

    def set_eye_event(self,eye_event):
        if len(self.eye_event_queue) < self.cfg_db['eye_queue_max']:
            if type(eye_event) is list:
                for event in eye_event:
                    self.eye_event_queue.append(event)
            else:
                self.eye_event_queue.append(eye_event)                
        
    def handle_events(self,events):
        if len(events) == 0:
            return

        now_sec = time.time()        
        for event in events:
            if False:
                print ('handle_event: {}'.format(event))
            if type(event) is tuple:
                self.set_eye_event(event)
            elif type(event) is list:
                self.set_eye_event(event)                
            elif event in ['pupil_widen','pupil_narrow']:
                if not self.pupil_event_queued:
                    self.event_pupil = event
                    self.pupil_event_queued = True
            elif event in ['blink']:
                self.event_blink ^= 1
                self.event_doBlink = True
                if self.cfg_db['joystick_mode'] in [1]:
                    self.isMoving = False
                self.event_overrideBlinkDurationClose = \
                    random.uniform(self.cfg_db['blink_duration_joystickmin_sec'],
                                   self.cfg_db['blink_duration_joystickmax_sec'])
	        self.timeOfLastBlink = now_sec
                self.override_timeToNextBlink = self.event_overrideBlinkDurationClose + \
                    self.cfg_db['blink_interval_min_sec'] + \
                    random.uniform(0.0,self.cfg_db['blink_interval_range_sec'])
                
            elif event in ['eye_context_9'] and self.cfg_db['demo']:
                self.eye_context_next = 'dragon'
            elif event in ['eye_context_11'] and self.cfg_db['demo']:
                self.eye_context_next = 'cyclops'
            elif event in ['eye_context_12'] and self.cfg_db['demo']:
                self.eye_context_next = 'hack'
            elif event in ['eye_goto']:
                raise
            else:
                print ('** Unhandled event: {}'.format(event))
                continue
            
        self.update_eye_events()

    def random_next_eye(self):
        next_eye = self.EYE_SELECT
        while next_eye is self.EYE_SELECT:
            if self.cfg_db['playa']:
                break
            else:
                #next_eye = random.choice(['cyclops','dragon','hack'])
                next_eye = random.choice(self.eye_contexts)

        return next_eye
        
    def map_wearables_events(self,wearable_events):
        eye_events = []
        for event in wearable_events:
            if event in ['slowblink']:
                eye_events.append('eye_up')
            elif event in ['radiaterainbow']:
                eye_events.append('eye_northeast')
            elif event in ['rider']:
                eye_events.append('eye_right')
            elif event in ['threesine']:
                eye_events.append('eye_southeast')
            elif event in ['flame']:
                eye_events.append('eye_down')
                eye_events.append('eye_southwest')                
            elif event in ['glitter']:
                eye_events.append('eye_left')
                eye_events.append('eye_northwest')
            elif event in ['eye_goto']:
                raise
            else:
                print ('** Unmapped wearables event: {}'.format(event))

        if len(eye_events) > 0:
            next_eye = self.random_next_eye()
            if self.cfg_db['switch_on_blink']:
                if not self.eye_switch_pending:
                    self.eye_switch_pending = True
                    self.eye_switch_next = next_eye
            else:
                self.EYE_SELECT = self.switch_eye_context(next_eye)
                
        return eye_events
    
    def do_wearables(self):
        msgs = self.wearables_client.get_msgs_nonblocking()
        if msgs is not None:
            wearable_events = [msg_rec['effect'] for msg_rec in msgs]
            print ('wearable_events: {}'.format(wearable_events))
            self.wearables_msg_cnt += len(wearable_events)
            wearable_eye_events = self.map_wearables_events(wearable_events)
            print ('wearable_eye_events: {}'.format(wearable_eye_events))
            self.handle_events(wearable_eye_events)

    def sanity_check_comm(self,events):
        if self.eye_comm_msg_cnt == 0:
            self.debug_idx = self.test_eye_events.index(events[0])

        for idx,event in enumerate(events):
            if event != self.test_eye_events[(self.debug_idx + idx) % \
                                             len(self.test_eye_events)]:
                print ('** Error: {} is not expected: {}'.format(
                    event,
                    self.test_eye_events[(self.debug_idx + idx) % \
                                         len(self.test_eye_events)]))
                raise
            else:
                print ('event: {} is ok'.format(event))

        self.debug_idx += len(events)                    
        
    def do_eye_comm(self):
        msgs = self.eye_client.get_msgs_nonblocking()
        if msgs is not None:
            self.last_eye_comm_recv = time.time()
            gecko_events = [msg_rec['effect'] for msg_rec in msgs]
            if self.debug:
                print ('eye_comm: recv {}'.format(gecko_events))
            #self.sanity_check_comm(gecko_events)
            self.eye_comm_msg_cnt += len(gecko_events)
            if self.debug:
                print ('eye_comm_msg_cnt: {}'.format(self.eye_comm_msg_cnt))
                
            if self.joystick is None:
                self.handle_events(gecko_events)
            else:
                print ('Joystick connected master ignoring loopback events')
            
    def create_joystick_test_msg(self):
        #eye_event = random.choice(self.test_eye_events)
        eye_event = self.test_eye_events[self.test_joystick_cnt % len(self.test_eye_events)]
        now = time.time()
        dt = now - self.debug_joystick_sec
        if dt > 1:
            self.eye_server.send_msg(eye_event)
            self.debug_joystick_sec = now
            self.test_joystick_cnt += 1
        
    def do_joystick(self):
        # self.init_joystick()
        gecko_events = []
        if self.cfg_db['joystick_test']:
            gecko_events = self.create_joystick_test_msg()
        elif self.joystick is not None:
            #print ('DO_JOYSTICK')
            gecko_events = self.joystick.sample_nonblocking()
            self.handle_events(gecko_events)
            eye_event_last = None
            for eye_event in gecko_events:
                # Debounce certain events
                if eye_event is eye_event_last and \
                   eye_event in ['blink']:
                    continue

                if self.cfg_db['demo']: # some day also or self.cfg_db['playa']
                    print ('eye_comm: send: {}'.format(eye_event))
                    if type(eye_event) is list:
                        for event in eye_event:
                            self.eye_server.send_msg(event)
                        eye_event_last = event
                    else:
                        self.eye_server.send_msg(eye_event)
                        eye_event_last = eye_event
            self.joystick_polls +=1
            self.joystick_msg_cnt += len(gecko_events)
        
        if self.debug:
            print ('joystick_polls: {}'.format(self.joystick_polls))

        if gecko_events is None:
            return 0
        
        return (len(gecko_events) > 0)
    

    def screenshots(self):
	self.DISPLAY.loop_running()        
	# Draw eye
        for fname_sclera in self.hack_scleras:
            for fname_iris in self.hack_iris:
                # create a special eye context
                screenshot_context = 'screenshot'
                self.cfg_db[screenshot_context] = {
		    'eye.shape': 'graphics/cyclops-eye.svg',
		    'iris.art': fname_iris,
		    'lid.art': 'graphics/lid.png',
		    'sclera.art': fname_sclera
		}
                self.init(['screenshot'])
                self.EYE_SELECT = self.switch_eye_context(screenshot_context)
                
                for i in range(2):
                    self.frame(1.0)

                # Sclera
                re_fname = re.compile(r'^(.*)\.(\S+)$',re.IGNORECASE)
                fname = fname_sclera.split('/')[-1]
                print ('fname_sclera: {}'.format(fname))
                m = re_fname.match(fname)
                assert (m)
                print ('m: {}'.format(m.groups()))
                fname_base_sclera= m.group(1)

                # Iris
                fname = fname_iris.split('/')[-1]
                print ('fname_iris: {}'.format(fname))
                m = re_fname.match(fname_iris.split('/')[-1])
                assert (m)
                print ('m: {}'.format(m.groups()))                
                fname_base_iris = m.group(1)
                
                fname_screenshot = '{}_{}.jpg'.format(fname_base_sclera,fname_base_iris)
                fname_screenshot.replace(' ','_')
                print ('fname_screenshot: {}'.format(fname_screenshot))
                pi3d.screenshot(fname_screenshot)
                #time.sleep(1)

    def check_settings(self):
        POLL_INTERVAL_SEC = 1
        cfg_db = {'key' : 'old_value'}
        time_now = datetime.datetime.now()
        time_delta = time_now - self.time_last_poll
        #print (time_delta)
        if time_delta.seconds > POLL_INTERVAL_SEC:
            self.time_last_poll = time_now
            changed = self.json_db.poll_json_changed()
            if changed:
                self.cfg_db_new = self.json_db.read_json_db(self.cfg_db)
                print ('******* NEW JSON DB ************')
                print (self.cfg_db_new)
                print ('******* END JSON DB ************')
                #self.eye_switch_pending = True
                self.cfg_db = self.cfg_db_new
                self.EYE_SELECT = self.switch_eye_context(self.EYE_SELECT)                
                
    def run(self):
        if self.cfg_db['screenshots']:
            self.screenshots()
            DISPLAY.destroy()
            sys.exit(0)
        
        do_exit = False
        if self.cfg_db['demo']:
            self.cfg_db['eye_orientation'] = random.choice(['left','right'])
        elif self.cfg_db['playa']:
            self.cfg_db['eye_orientation'] = 'right'
            
        print ('eye_orientation: {}'.format(self.cfg_db['eye_orientation']))

        now_time = time.time()
        self.run_start_time = now_time
        self.watchdog_sec = now_time
        self.last_eye_art_sec = now_time
        while not do_exit:

            # Check for settings changes
            self.check_settings()
            
            if self.cfg_db['PUPIL_IN'] >= 0: # Pupil scale from sensor
                raise
		v = adcValue[self.cfg_db['PUPIL_IN']]
		if self.cfg_db['PUPIL_IN_FLIP']: v = 1.0 - v
		# If you need to calibrate PUPIL_MIN and MAX,
		# add a 'print v' here for testing.
		if   v < self.cfg_db['PUPIL_MIN']: v = self.cfg_db['PUPIL_MIN']
		elif v > self.cfg_db['PUPIL_MAX']: v = self.cfg_db['PUPIL_MAX']
		# Scale to 0.0 to 1.0:
		v = (v - self.cfg_db['PUPIL_MIN']) / (self.cfg_db['PUPIL_MAX'] -
                                                      self.cfg_db['PUPIL_MIN'])
		if self.cfg_db['PUPIL_SMOOTH'] > 0:
			v = ((currentPupilScale * (self.cfg_db['PUPIL_SMOOTH'] - 1) + v) /
			     self.cfg_db['PUPIL_SMOOTH'])
                self.frame(v)
            else: # Fractal auto pupil scale
                # Priority mux
                if self.eye_context_next is not None: # Transition to new eye
                    break
                
                elif self.pupil_event_queued:  # Joystick control of pupil
                    self.pupil_event_queued = False
                    self.eye_event_prev = self.event_pupil
                    if self.event_pupil in ['pupil_widen']:
                        v = 1.0
                    elif self.event_pupil in ['pupil_narrow']:
                        v = 0.0
                    else:
                        raise
                    duration = 0.25
                else:
                    emotion_selected = self.emotion_select()
                    emotion_selected = None
                    if emotion_selected is not None:
                        (v,duration) = emotion_selected()
                        if v is None:
                            v = self.last_v
                            duration = 0.0
                        self.last_v = v
                    else: # Autonomous mode
                        v = self.last_v
                        #duration = 0.1
                        #v = random.random()
                        duration = self.cfg_db['pupil_auto_expand_sec']
                do_exit |= self.split(self.currentPupilScale, v, duration, 1.0)
                #leak_check()                
                

            self.currentPupilScale = v
            #do_exit = self.keyboard_sample()
            now_sec = time.time()
            if self.cfg_db['timeout_secs'] is not None and \
               int(now_sec - self.watchdog_sec) > self.cfg_db['timeout_secs']:
                do_exit |= True

        if do_exit:
            print ('exiting')
            if self.cfg_db['demo']:
                self.eye_context_next = self.EYE_SELECT
            elif self.EYE_SELECT in ['hack']:
                self.eye_context_next = self.EYE_SELECT                                
            else:
                self.eye_context_next = None

        return self.eye_context_next

    def shutdown(self):
        #del self.light
        #del self.cam
        #del self.shader
        #self.joystick.shutdown()
        del self.parser
        self.parser = None
        pass

    def init_emotion(self):
        self.emotions = [
            self.emotion_normal,
            self.emotion_random_looking,
            self.emotion_dilated_pupil,
            self.emotion_squint_pupil,
            self.emotion_wandering,
            self.emotion_angry,
            self.emotion_staring_ahead
        ]
        self.emotions = [None]
        self.emotions = [self.emotion_squint_pupil]
        self.emotions = [self.emotion_dilated_pupil]
        self.emotions = [
            self.emotion_dilated_pupil,
            self.emotion_squint_pupil,
            self.emotion_normal,
            self.emotion_staring_ahead            
        ]
        self.emotions = [self.emotion_staring_ahead]
        self.emotions = [self.emotion_angry]
        
        self.emotion_idx = 0
        self.shuffled_emo = self.emotions
        self.fsm_angry = None
        
    def emotion_select(self):
        now_sec = time.time()
        if now_sec < self.nxt_emotion_sec:
            return None

        if self.emotion_idx == 0:
            random.shuffle(self.emotions)

        self.fsm_angry = None
        selected_emotion = self.shuffled_emo[self.emotion_idx]

        if self.fsm_angry is None:
            self.nxt_emotion_sec = now_sec + self.cfg_db['emotion_interval_sec']
        else:
            self.nxt_emotion_sec = self.cfg_db['emotion_interval_sec'] / 4
        
        self.emotion_idx += 1
        self.emotion_idx %= len(self.emotions)

        return selected_emotion
    
    def emotion_normal(self):
        print ('emotion_normal')
        v = self.cfg_db['pupil_normal']
        duration = self.cfg_db['pupil_dilate_sec']
        return (v,duration)        

    def emotion_random_looking(self):
        print ('emotion_random_looking')
        return (None,None)        

    def emotion_dilated_pupil(self):
        print ('emotion_dilated_pupil')
        v = self.cfg_db['pupil_max']
        duration = self.cfg_db['pupil_dilate_sec']
        return (v,duration)        
    
    def emotion_squint_pupil(self):
        print ('emotion_squint_pupil')
        v = self.cfg_db['pupil_min']
        duration = self.cfg_db['pupil_squint_sec']
        return (v,duration)

    def emotion_wandering(self):
        print ('emotion_wandering')
        return (None,None)

    def emotion_angry(self):
        print ('emotion_angry')
        if self.fsm_angry is None:
            self.fsm_states = ['focused','alarmed']
            self.fsm_angry = random.choice(self.fsm_states)

        if self.fsm_angry in ['focused']:
            v = self.cfg_db['pupil_min']
            duration = self.cfg_db['pupil_squint_sec']
            
            focused_dirs = [
                'eye_southwest',
                'eye_southeast',
                'eye_down'
            ]
            
            self.eye_event = random.choice(focused_dirs)
            print ('angry dir: {}'.format(self.eye_event))            
            #self.event_moveDuration = self.cfg_db['move_angry_duration_min_sec']
            self.event_moveDuration = 0.25
            self.event_holdDuration = 2.0
            self.event_overrideBlinkDurationClose = \
                random.uniform(self.cfg_db['blink_angry_duration_close_min_sec'],
                               self.cfg_db['blink_angry_duration_close_max_sec'])
        
            self.event_overrideBlinkDurationOpen = \
                random.uniform(self.cfg_db['blink_angry_duration_open_min_sec'],
                               self.cfg_db['blink_angry_duration_open_max_sec'])
            self.event_doBlink = True            
        elif self.fsm_angry in ['alarmed']:
            v = self.cfg_db['pupil_max']
            duration = self.cfg_db['pupil_dilate_sec']
            self.event_overrideBlinkDurationClose = \
                random.uniform(self.cfg_db['blink_angry_duration_close_min_sec'],
                               self.cfg_db['blink_angry_duration_close_max_sec'])
        
            self.event_overrideBlinkDurationOpen = \
                random.uniform(self.cfg_db['blink_angry_duration_open_min_sec'],
                               self.cfg_db['blink_angry_duration_open_max_sec'])
            self.event_moveDuration = random.uniform(self.cfg_db['move_angry_duration_min_sec'],
                                                     self.cfg_db['move_angry_duration_max_sec'])
        
            self.event_holdDuration = random.uniform(self.cfg_db['hold_angry_duration_min_sec'],
                                                     self.cfg_db['hold_angry_duration_max_sec'])
            self.event_doBlink = True
        else:
            raise

        self.fsm_states = ['focused','alarmed']
        self.fsm_angry_nxt = random.choice(self.fsm_states)
        self.fsm_angry = self.fsm_angry_nxt
        print ('fsm_angry: {}'.format(self.fsm_angry))
        
        #return (None,None)        
        
        return (v,duration)
    

    def emotion_staring_ahead(self):
        print ('emotion_staring_ahead')
        staring_head_dirs = [
            'eye_northeast',
            'eye_southeast',
            'eye_right'
        ]
        
        self.eye_event = random.choice(staring_head_dirs)
        print ('staring head dir: {}'.format(self.eye_event))
        self.event_moveDuration = self.cfg_db['move_duration_min_sec']
        self.event_holdDuration = 20.0
        
        return (None,None)        

        
if __name__ == "__main__":
    eye_context = None
    time_first_sec = time.time()    
    timeout = False
    while True and not timeout:
        leak_check()
        gecko_eye = gecko_eye_t(EYE_SELECT=eye_context)
        eye_context = gecko_eye.run()
        if gecko_eye.cfg_db['timeout_secs'] is not None:
            now_sec = time.time()
            if now_sec > time_first_sec + gecko_eye.cfg_db['timeout_secs']:
                timeout = True
            
        gecko_eye.shutdown()
        if eye_context is None:
            break

    DISPLAY.destroy()
        
    sys.exit(0)
